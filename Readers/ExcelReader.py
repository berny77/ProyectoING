from openpyxl import load_workbook

class ExcelReader:
    def __init__(self, ruta):
        self.ruta = ruta

    def leer(self):
        try:
            contenido = ""
            wb = load_workbook(self.ruta, data_only=True)
            for hoja in wb.sheetnames:
                ws = wb[hoja]
                contenido += f"--- Hoja: {hoja} ---\n"
                for fila in ws.iter_rows(values_only=True):
                    fila_texto = "\t".join([str(celda) if celda is not None else "" for celda in fila])
                    contenido += fila_texto + "\n"
            return contenido
        except Exception as e:
            raise ValueError(f"Error al procesar el archivo Excel: {e}")
